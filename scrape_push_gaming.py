#!/usr/bin/env python3
from __future__ import annotations
"""
Scrape Push Gaming slot pages:
  Task 1 — fill missing RTP, volatility, highest_win, bonus_buy, features
  Task 2 — download thumbnail and hero images for all 85 games
"""

import json
import re
import time
import urllib.request
from pathlib import Path

import openpyxl
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout

XLSX_PATH = Path(__file__).parent / "Features_push_gaming.xlsx"
IMAGES_DIR = Path(__file__).parent / "images" / "push_gaming"
IMAGES_DIR.mkdir(parents=True, exist_ok=True)

FAILED_DATA: list[str] = []
FAILED_IMAGES: list[str] = []
BASE_URL = "https://www.pushgaming.com"


def dismiss_age_gate(page):
    try:
        btn = page.locator("text='I am over 18'").first
        btn.wait_for(timeout=3000)
        btn.click()
        page.wait_for_timeout(800)
    except PWTimeout:
        pass


def goto(page, url: str):
    page.goto(url, wait_until="domcontentloaded", timeout=30000)
    dismiss_age_gate(page)
    page.wait_for_timeout(1500)


def scrape_game_data(page, url: str, name: str) -> dict | None:
    try:
        goto(page, url)
        content = page.content()

        # RTP — appears as "96.55%<br>94.49%" in .game-features
        rtps = list(dict.fromkeys(re.findall(r'\b(9[0-9]\.\d{1,2})%', content)))
        rtp_base    = rtps[0] + '%' if len(rtps) >= 1 else ''
        rtp_reduced = rtps[1] + '%' if len(rtps) >= 2 else ''

        # Volatility — from icon img src: Volatility_High.png
        vol_match = re.search(r'Volatility_(\w+)\.png', content)
        volatility = vol_match.group(1).replace('_', ' ') if vol_match else ''

        # Highest win — from .game-features: "100,000x" before RTPS/VOLATILITY block
        win_match = re.search(
            r'HIGHEST[^<]*</p>\s*<p[^>]*>\s*([\d,]+(?:\.\d+)?x?)',
            content, re.IGNORECASE
        )
        if not win_match:
            # broader fallback: number+x in the stats block text
            stats_text = page.locator('.game-features').first.inner_text(timeout=3000)
            wm = re.search(r'([\d,]+(?:\.\d+)?)x', stats_text)
            win_val = wm.group(1) if wm else ''
        else:
            win_val = win_match.group(1).strip()
        highest_win = (win_val + 'x') if win_val and not win_val.endswith('x') else win_val

        # Bonus Buy
        bonus_buy = 'Yes' if re.search(r'BONUS[\s-]*BUY', content, re.IGNORECASE) else 'No'

        # Features — from .img-text-grid (deduplicated)
        features_raw = page.evaluate("""
            () => {
                const seen = new Set();
                const results = [];
                for (const el of document.querySelectorAll('.img-text-grid')) {
                    const lines = el.innerText.trim()
                        .split('\\n')
                        .map(s => s.trim())
                        .filter(Boolean);
                    if (lines.length < 2) continue;
                    const title = lines[0];
                    if (seen.has(title)) continue;
                    seen.add(title);
                    const description = lines.slice(1).join(' ').replace(/\\s+/g, ' ').trim();
                    results.push({title, description});
                }
                return results;
            }
        """)
        features = json.dumps(features_raw) if features_raw else ''

        return {
            'rtp_base':    rtp_base,
            'rtp_reduced': rtp_reduced,
            'volatility':  volatility,
            'highest_win': highest_win,
            'bonus_buy':   bonus_buy,
            'features':    features,
        }

    except Exception as e:
        print(f"    ERROR scraping {name}: {e}")
        return None


def fetch_all_thumbnails(page) -> dict[str, str]:
    """Load the games listing page once and return {slug: thumbnail_url} for all 85 games."""
    page.goto(f"{BASE_URL}/games", wait_until="domcontentloaded", timeout=30000)
    dismiss_age_gate(page)
    page.wait_for_timeout(1000)
    # Scroll to ensure lazy-loaded cards are rendered
    for _ in range(12):
        page.evaluate("window.scrollBy(0, 800)")
        page.wait_for_timeout(250)

    return page.evaluate("""
        () => {
            const result = {};
            for (const link of document.querySelectorAll('a[href*="/games/"]')) {
                const m = link.href.match(/\\/games\\/([^.]+)\\.html/);
                if (!m) continue;
                const slug = m[1];
                if (result[slug]) continue;
                let el = link;
                for (let i = 0; i < 6; i++) {
                    if (!el) break;
                    const targets = [el, ...el.querySelectorAll('.bg,[class*="bg"]')];
                    for (const t of targets) {
                        const bg = window.getComputedStyle(t).backgroundImage;
                        if (bg && bg !== 'none') {
                            const m2 = bg.match(/url\\(["']?(.*?)["']?\\)/);
                            if (m2 && m2[1] && !m2[1].includes('sub-box')) {
                                result[slug] = m2[1];
                                break;
                            }
                        }
                    }
                    if (result[slug]) break;
                    el = el.parentElement;
                }
            }
            return result;
        }
    """)


def get_hero_image(page, url: str) -> str | None:
    """Navigate to the game page and return its hero/background image URL."""
    try:
        goto(page, url)
        return page.evaluate("""
            () => {
                for (const el of document.querySelectorAll('[class*="responsimg"], [class*="hero"], [class*="banner"]')) {
                    const bg = window.getComputedStyle(el).backgroundImage;
                    if (bg && bg !== 'none') {
                        const m = bg.match(/url\\(["']?(.*?)["']?\\)/);
                        if (m && m[1] && !m[1].includes('sub-box')) return m[1];
                    }
                }
                return null;
            }
        """)
    except Exception:
        return None


def download_image(url: str, dest: Path) -> bool:
    try:
        if url.startswith('//'):
            url = 'https:' + url
        if url.startswith('/'):
            url = BASE_URL + url
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=20) as resp:
            data = resp.read()
        dest.write_bytes(data)
        return True
    except Exception as e:
        print(f"    Download failed {url}: {e}")
        return False


def main():
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    col = {h: i + 1 for i, h in enumerate(headers)}

    all_rows = []
    for row_idx in range(2, ws.max_row + 1):
        row_data = {h: ws.cell(row_idx, col[h]).value for h in headers}
        row_data['_row_idx'] = row_idx
        all_rows.append(row_data)

    missing_data_rows = [r for r in all_rows if not r.get('rtp_base')]
    print(f"Rows needing data: {len(missing_data_rows)} / {len(all_rows)}")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        ctx = browser.new_context(
            viewport={"width": 1280, "height": 900},
            user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/120 Safari/537.36"
        )
        page = ctx.new_page()

        # ── Task 1: Fill missing data ─────────────────────────────────────────
        if missing_data_rows:
            print("\n=== TASK 1: Scraping missing game data ===")
            for row in missing_data_rows:
                name = row['name']
                print(f"  {name} ... ", end='', flush=True)
                data = scrape_game_data(page, row['pushgaming_url'], name)
                if data:
                    for field in ['rtp_base', 'rtp_reduced', 'volatility', 'highest_win', 'bonus_buy', 'features']:
                        if data.get(field):
                            ws.cell(row['_row_idx'], col[field], data[field])
                    feat_count = len(json.loads(data['features'])) if data.get('features') else 0
                    print(f"RTP:{data['rtp_base']} Vol:{data['volatility']} Win:{data['highest_win']} BB:{data['bonus_buy']} F:{feat_count}")
                else:
                    FAILED_DATA.append(name)
                    print("FAILED")
                time.sleep(0.3)

            wb.save(XLSX_PATH)
            print(f"Spreadsheet saved. Failures: {FAILED_DATA or 'none'}")

        # ── Task 2: Download images ───────────────────────────────────────────
        print("\n=== TASK 2: Downloading images ===")
        print("  Fetching all thumbnails from listing page...")
        thumbnails = fetch_all_thumbnails(page)
        print(f"  Got {len(thumbnails)} thumbnails")

        for row in all_rows:
            name = row['name']
            slug = row['slug']
            print(f"  {name}")

            # Thumbnail
            thumb_url = thumbnails.get(slug)
            if thumb_url:
                ext = ('.' + thumb_url.split('?')[0].rsplit('.', 1)[-1]) if '.' in thumb_url.rsplit('/', 1)[-1] else '.jpg'
                dest = IMAGES_DIR / f"{slug}{ext}"
                ok = download_image(thumb_url, dest)
                print(f"    thumb -> {dest.name}" if ok else f"    thumb FAILED: {thumb_url}")
                if not ok:
                    FAILED_IMAGES.append(f"{name} (thumb dl)")
            else:
                print(f"    thumb: not found on listing")
                FAILED_IMAGES.append(f"{name} (thumb missing)")

            # Hero from game page
            hero_url = get_hero_image(page, row['pushgaming_url'])
            if hero_url:
                ext = ('.' + hero_url.split('?')[0].rsplit('.', 1)[-1]) if '.' in hero_url.rsplit('/', 1)[-1] else '.jpg'
                dest = IMAGES_DIR / f"{slug}_hero{ext}"
                ok = download_image(hero_url, dest)
                print(f"    hero  -> {dest.name}" if ok else f"    hero  FAILED: {hero_url}")
                if not ok:
                    FAILED_IMAGES.append(f"{name} (hero dl)")
            else:
                print(f"    hero: not found")
                FAILED_IMAGES.append(f"{name} (hero missing)")

            time.sleep(0.2)

        browser.close()

    print("\n=== DONE ===")
    print(f"Data failures  : {FAILED_DATA or 'none'}")
    print(f"Image failures : {FAILED_IMAGES or 'none'}")
    print(f"Images saved to: {IMAGES_DIR}")


if __name__ == '__main__':
    main()
